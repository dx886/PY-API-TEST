# 2019/04/13 完成do_Excel类的封装
# 1，完成do_excel类的封装支持读和写，理解Case类的设计
## 2，结合http_requests完成注册，登录，充值接口的请求
# 3，尝试引入unittest+ddt 来完成以上接口的测试用例、
from openpyxl import load_workbook
from python_study.zuoye.zuoye_17.zuoye_request_17 import HttpRequest
class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]
    def get_cases(self):
        max_row = self.sheet.max_row  # 获取最大行数
        cases = []  # 列表，存放所有的测试用例
        for r in range(2, max_row + 1):
            case = Case()
            case.case_id = self.sheet.cell(row=r, column=2).value
            case.title = self.sheet.cell(row=r, column=3).value
            case.url = self.sheet.cell(row=r, column=4).value
            case.data = self.sheet.cell(row=r, column=5).value
            case.method = self.sheet.cell(row=r, column=6).value
            case.expected = self.sheet.cell(row=r, column=7).value
            cases.append(case)
        self.workbook.close()
        return cases  # 返回case列表
    def write_result(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 8).value=actual
        sheet.cell(row, 9).value=result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()

if __name__ == '__main__':

     excel=DoExcel('cases.xlsx', 'login')
     excel1=DoExcel('cases.xlsx', 'login').get_cases()
     request1 = HttpRequest()
     for i in excel1:
            print(i.__dict__)
     #        resp = request1.Request(i.method,i.url,eval(i.data))
     #        print(resp.text)
     #        actual = resp.text
     #        if i.expected == actual:
     #            excel.write_result(i.case_id + 1, actual, 'PASS')
     #
     #        else:
     #            excel.write_result(i.case_id + 1, actual, 'FAIL')